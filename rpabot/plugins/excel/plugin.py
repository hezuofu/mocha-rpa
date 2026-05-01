"""Excel plugin for the RPA framework.

Provides Excel read/write operations via ``openpyxl``, integrated with the
framework's plugin lifecycle.

Usage::

    from rpabot.plugins.excel.plugin import ExcelPlugin
    from rpabot.plugin.base import PluginManager

    mgr = PluginManager(context)
    excel = ExcelPlugin()
    mgr.register(excel)
    mgr.start_all()

    wb = excel.open("report.xlsx")
    excel.write_cell(wb, "Sheet1", 1, 1, "Hello")
    excel.save("report.xlsx")
    mgr.shutdown_all()
"""

from __future__ import annotations

import logging
from contextlib import contextmanager
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from rpabot.plugin.base import Plugin

logger = logging.getLogger("rpa.excel")


class ExcelPlugin:
    """Plugin for Excel workbook operations using ``openpyxl``.

    Attributes:
        name: Plugin identifier (``"excel"``).
    """

    name = "excel"

    def __init__(self) -> None:
        self._workbooks: Dict[str, Workbook] = {}
        self._context: Any = None

    # ------------------------------------------------------------------
    # Plugin lifecycle
    # ------------------------------------------------------------------

    def initialize(self, context: Any) -> None:
        self._context = context
        logger.info("ExcelPlugin initialized")

    def cleanup(self) -> None:
        for name in list(self._workbooks):
            self.close(name)
        logger.info("ExcelPlugin cleaned up")

    # ------------------------------------------------------------------
    # Workbook management
    # ------------------------------------------------------------------

    def open(self, path: str) -> Workbook:
        """Open an existing workbook.  Cached by filename stem.

        Returns:
            The :class:`openpyxl.Workbook` instance.
        """
        name = self._key_from_path(path)
        if name in self._workbooks:
            return self._workbooks[name]
        wb = openpyxl.load_workbook(path)
        self._workbooks[name] = wb
        logger.info("Opened workbook: %s", path)
        return wb

    def create(self, path: str = "workbook.xlsx") -> Workbook:
        """Create a new workbook (or overwrite existing)."""
        name = self._key_from_path(path)
        wb = openpyxl.Workbook()
        self._workbooks[name] = wb
        logger.info("Created workbook: %s", path)
        return wb

    def save(self, name: str, path: Optional[str] = None) -> None:
        """Save a managed workbook.

        Args:
            name: Workbook key (filename stem or full path).
            path: Optional save target.  If ``None``, the key is used as filename.
        """
        wb = self._workbooks.get(name)
        if wb is None:
            raise KeyError(f"Workbook '{name}' not found")
        target = path or f"{name}.xlsx"
        wb.save(target)
        logger.info("Saved workbook: %s", target)

    def close(self, name: str, *, save: bool = False) -> None:
        """Close a managed workbook.

        Args:
            name: Workbook key.
            save: If True, save before closing.
        """
        wb = self._workbooks.pop(name, None)
        if wb is None:
            return
        if save:
            wb.save(name if "." in name else f"{name}.xlsx")
        wb.close()
        logger.info("Closed workbook: %s", name)

    def get(self, name: str) -> Workbook:
        """Return a managed workbook by name."""
        wb = self._workbooks.get(name)
        if wb is None:
            raise KeyError(f"Workbook '{name}' not found. Use open() or create() first.")
        return wb

    def list_workbooks(self) -> List[str]:
        """Return names of all open workbooks."""
        return list(self._workbooks.keys())

    # ------------------------------------------------------------------
    # Sheet operations
    # ------------------------------------------------------------------

    @staticmethod
    def list_sheets(wb: Workbook) -> List[str]:
        """Return sheet names in *wb*."""
        return wb.sheetnames

    @staticmethod
    def add_sheet(wb: Workbook, name: str) -> Worksheet:
        """Add a new sheet."""
        return wb.create_sheet(title=name)

    @staticmethod
    def get_sheet(wb: Workbook, name: str) -> Worksheet:
        """Get a sheet by name."""
        if name in wb.sheetnames:
            return wb[name]
        raise KeyError(f"Sheet '{name}' not found in workbook")

    @staticmethod
    def remove_sheet(wb: Workbook, name: str) -> None:
        """Remove a sheet by name."""
        if name not in wb.sheetnames:
            raise KeyError(f"Sheet '{name}' not found")
        del wb[name]

    # ------------------------------------------------------------------
    # Cell operations
    # ------------------------------------------------------------------

    @staticmethod
    def read_cell(wb: Workbook, sheet: str, row: int, col: int) -> Any:
        """Read a single cell value.  Row/col are 1-indexed."""
        ws = wb[sheet]
        return ws.cell(row=row, column=col).value

    @staticmethod
    def write_cell(wb: Workbook, sheet: str, row: int, col: int, value: Any) -> None:
        """Write a value to a cell.  Row/col are 1-indexed."""
        ws = wb[sheet]
        ws.cell(row=row, column=col, value=value)

    # ------------------------------------------------------------------
    # Range operations
    # ------------------------------------------------------------------

    @staticmethod
    def read_range(
        wb: Workbook,
        sheet: str,
        start: Tuple[int, int],
        end: Tuple[int, int],
    ) -> List[List[Any]]:
        """Read a rectangular range as a list of rows.

        Args:
            start: ``(row, col)`` top-left, 1-indexed.
            end: ``(row, col)`` bottom-right, 1-indexed.
        """
        ws = wb[sheet]
        data: List[List[Any]] = []
        for r in range(start[0], end[0] + 1):
            row_data: List[Any] = []
            for c in range(start[1], end[1] + 1):
                row_data.append(ws.cell(row=r, column=c).value)
            data.append(row_data)
        return data

    @staticmethod
    def write_range(
        wb: Workbook,
        sheet: str,
        start_row: int,
        start_col: int,
        data: List[List[Any]],
    ) -> None:
        """Write a 2D list of values starting at *start_row*, *start_col* (1-indexed)."""
        ws = wb[sheet]
        for ri, row_data in enumerate(data):
            for ci, value in enumerate(row_data):
                ws.cell(row=start_row + ri, column=start_col + ci, value=value)

    @staticmethod
    def read_all(wb: Workbook, sheet: str) -> List[List[Any]]:
        """Read all data from a sheet as a list of rows."""
        ws = wb[sheet]
        return [[cell.value for cell in row] for row in ws.iter_rows()]

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _key_from_path(path: str) -> str:
        import os
        return os.path.splitext(os.path.basename(path))[0]
